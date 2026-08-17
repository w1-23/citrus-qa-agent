"""File reading tool — read_local_file (system-wide read, workspace-relative as fallback)

v8.3.0: max_chars parameter — default 30000 for PDF, unlimited for others.
v8.10r: PDF 默认读全文（max_chars<=0 = 不截断）——用户明确 PDF 读取不应有
默认阈值截断；如需限制仍可显式传 max_chars>0。
"""
import logging
import os
from pathlib import Path

from langchain_core.tools import tool

from src.config import settings, PROJECT_ROOT

logger = logging.getLogger(__name__)

_WORKSPACE_ROOT = PROJECT_ROOT / settings.WORKSPACE_DIR


def _is_path_allowed(full: Path) -> bool:
    """路径白名单判定（v8.13 收敛到 core.path_policy.is_within）。"""
    from src.core.path_policy import is_within, read_roots
    return is_within(full, read_roots())


def _resolve_read_path(path: str) -> Path:
    """路径解析（v8.13 收敛到 core.path_policy.resolve_read）。

    绝对路径仅允许项目根+额外根；相对路径锚定 workspace/（不再 Path.cwd()
    兜底，杜绝相对 + .. 逃逸）。越界抛 PermissionError（fail-closed）。
    """
    from src.core.path_policy import resolve_read
    try:
        full = resolve_read(path)
    except PermissionError:
        raise PermissionError(f"拒绝读取项目目录外的文件: {path}")
    if not full.exists():
        raise FileNotFoundError(f"文件不存在: {full}")
    return full


def _read_pdf(path: Path, max_chars: int = 0) -> str:
    """Read PDF pages. max_chars<=0 = 全文不截断；>0 时累计超过阈值停止。"""
    try:
        import fitz
        doc = fitz.open(path)
        texts = []
        total_chars = 0
        for i, page in enumerate(doc):
            text = page.get_text()
            if text and text.strip():
                page_block = f"--- Page {i+1} ---\n{text}"
                texts.append(page_block)
                total_chars += len(page_block)
                if max_chars > 0 and total_chars >= max_chars:
                    break
        total = len(doc)
        doc.close()
        valid = len(texts)
        if valid:
            header = f"[PDF: {path.name} | {valid} of {total} pages | fitz]\n\n"
            return header + "\n".join(texts)
        return f"[PDF: {path.name} | {total} pages | no extractable text]"
    except Exception:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            texts = []
            total_chars = 0
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    page_block = f"--- Page {i+1} ---\n{text}"
                    texts.append(page_block)
                    total_chars += len(page_block)
                    if max_chars > 0 and total_chars >= max_chars:
                        break
            total = len(pdf.pages)
            valid = len(texts)
            if valid:
                header = f"[PDF: {path.name} | {valid} of {total} pages | pdfplumber]\n\n"
                return header + "\n".join(texts)
            return f"[PDF: {path.name} | {total} pages | no extractable text]"
    except Exception:
        pass
    return f"[PDF: {path.name} | failed to read with both fitz and pdfplumber]"


def _read_excel(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        parts.append(f"## Sheet: {sheet_name} ({len(rows)} 行)")
        for i, row in enumerate(rows[:50]):
            parts.append(f"  Row {i+1}: {[str(c)[:60] if c else '' for c in row]}")
        if len(rows) > 50:
            parts.append(f"  ... (省略 {len(rows) - 50} 行)")
    wb.close()
    return "\n".join(parts)


def _read_csv(path: Path) -> str:
    import pandas as pd
    total_rows = sum(1 for _ in open(path, encoding="utf-8", errors="replace")) - 1
    df = pd.read_csv(path, nrows=1000)
    preview_json = df.to_json(orient="records", force_ascii=False)
    lines = [
        f"[CSV Preview] Total rows: {total_rows}. Showing first {len(df)} rows.\n",
        preview_json,
        f"\n\n*** Full dataset ({total_rows} rows) available in backend. "
        f"Use statistical_analysis(file_path='{path.name}') for full computations. ***",
    ]
    return "\n".join(lines)


def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    encodings = ["utf-8", "gbk", "gb2312", "latin-1", "shift_jis"]
    try:
        import chardet
        detected = chardet.detect(raw)
        enc = detected.get("encoding", "utf-8") or "utf-8"
        encodings.insert(0, enc)
    except ImportError:
        pass
    for enc in encodings:
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("latin-1", errors="replace")


@tool
def read_local_file(path: str, max_chars: int = 0) -> str:
    """读取本地文件。PDF 默认读全文（max_chars<=0 不截断；传正数可限制字符数）。

    绝对路径：仅允许项目根目录内的文件。
    相对路径：从 workspace/ 查找。

    v8.13 第四批: async def → def（A-06）——函数体全为同步 CPU/IO
    （fitz/pdfplumber/openpyxl/pandas），此前声明 async 导致 registry 直接
    `await tool.ainvoke` 挂在事件循环，50MB PDF 解析期间冻结所有会话。
    改为 sync 工具后统一经 executor 执行，不再阻塞事件循环。

    Args:
        path: 文件的绝对路径或相对路径
        max_chars: 最大读取字符数（<=0 = 不限制全文；>0 = 超过即停止）
    """
    logger.info(f"[read_local_file] path={path} max_chars={max_chars}")

    if not path or not path.strip():
        return "[ERR_PARSE] 文件路径不能为空"

    try:
        full_path = _resolve_read_path(path)
    except FileNotFoundError as e:
        return str(e)
    except Exception as e:
        return f"[ERR_PARSE] 路径解析失败: {e}"

    if not full_path.exists():
        return f"[ERR_FILE_NOT_FOUND] 文件不存在: {full_path}"

    size_mb = full_path.stat().st_size / (1024 * 1024)
    if size_mb > settings.FILE_READ_MAX_SIZE_MB:
        return f"[ERR_FILE_TOO_LARGE] 文件过大: {size_mb:.1f}MB > {settings.FILE_READ_MAX_SIZE_MB}MB 限制"

    ext = full_path.suffix.lower()

    try:
        if ext == ".pdf":
            content = _read_pdf(full_path, max_chars)
        elif ext in (".xlsx", ".xls"):
            content = _read_excel(full_path)
        elif ext == ".csv":
            content = _read_csv(full_path)
        else:
            content = _read_text(full_path)
        return content
    except Exception as e:
        logger.error(f"[read_local_file] 读取失败: {e}")
        return f"读取失败: {e}"
